from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

DAYS = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']
DAY_NAMES = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

day_pack = {
    'MON': [('Almonds', 6), ('Walnuts', 6), ('Dry Dates', 5), ('Salted Pista', 5), ('Pumpkin Seeds', 4), ('Yellow Raisins', 4)],
    'TUE': [('Almonds', 6), ('Walnuts', 6), ('Dry Dates', 5), ('Cashews', 5), ('Sunflower Seeds', 4), ('Black Raisins', 4)],
    'WED': [('Almonds', 6), ('Walnuts', 6), ('Dry Dates', 5), ('Salted Pista', 5), ('Pumpkin Seeds', 4), ('Yellow Raisins', 4)],
    'THU': [('Almonds', 6), ('Walnuts', 6), ('Dry Dates', 5), ('Cashews', 5), ('Sunflower Seeds', 4), ('Black Raisins', 4)],
    'FRI': [('Almonds', 6), ('Walnuts', 6), ('Pumpkin Seeds', 5), ('Cashews', 5), ('Yellow Raisins', 4), ('Black Raisins', 4)],
    'SAT': [('Almonds', 6), ('Walnuts', 6), ('Dry Dates', 5), ('Salted Pista', 5), ('Pumpkin Seeds', 4), ('Yellow Raisins', 4)],
    'SUN': [('Almonds', 6), ('Walnuts', 6), ('Dry Dates', 5), ('Salted Pista', 5), ('Sunflower Seeds', 4), ('Yellow Raisins', 4)],
}

night_soak = [
    ('Almond', 4), ('Walnut', 4), ('Cashew', 4), ('Salted Pista', 4),
    ('Black Raisins', 4), ('Pumpkin Seeds', 4), ('Sunflower Seeds', 4),
    ('Dry Figs', 4), ('Flax Seeds', 4), ('Chia Seed', 4), ('Black Dates', 4),
]

header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='2E4057')
day_header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
day_header_fill = PatternFill('solid', fgColor='048A81')
sub_header_font = Font(name='Arial', bold=True, size=10)
sub_header_fill = PatternFill('solid', fgColor='D6E4F0')
data_font = Font(name='Arial', size=10)
total_font = Font(name='Arial', bold=True, size=10)
total_fill = PatternFill('solid', fgColor='FFF2CC')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_cell(cell, font=None, fill=None, align=None, border=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border

def create_day_pack_sheet(ws):
    ws.title = 'Day Pack'
    ws.sheet_properties.tabColor = '048A81'
    sachet_weight = 30

    ws.merge_cells('A1:C1')
    c = ws['A1']
    c.value = 'Day Pack - Weekly Recipe (30g per sachet)'
    style_cell(c, Font(name='Arial', bold=True, color='FFFFFF', size=14), header_fill, Alignment(horizontal='center'))
    for col in range(1, 4):
        style_cell(ws.cell(row=1, column=col), fill=header_fill, border=thin_border)

    row = 3
    for i, day_key in enumerate(DAYS):
        ingredients = day_pack[day_key]
        ws.merge_cells(f'A{row}:C{row}')
        c = ws.cell(row=row, column=1, value=DAY_NAMES[i])
        style_cell(c, day_header_font, day_header_fill, Alignment(horizontal='center'))
        for col in range(1, 4):
            style_cell(ws.cell(row=row, column=col), fill=day_header_fill, border=thin_border)
        row += 1

        for label, col_idx in [('Ingredient', 1), ('Grams', 2), ('Percentage', 3)]:
            c = ws.cell(row=row, column=col_idx, value=label)
            style_cell(c, sub_header_font, sub_header_fill, Alignment(horizontal='center'), thin_border)
        row += 1

        for name, grams in ingredients:
            ws.cell(row=row, column=1, value=name).font = data_font
            ws.cell(row=row, column=2, value=grams).font = data_font
            ws.cell(row=row, column=2).number_format = '0.0'
            pct = round(grams / sachet_weight * 100, 1)
            ws.cell(row=row, column=3, value=pct / 100).font = data_font
            ws.cell(row=row, column=3).number_format = '0.0%'
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center' if col > 1 else 'left')
            row += 1

        c = ws.cell(row=row, column=1, value='Total')
        style_cell(c, total_font, total_fill, border=thin_border)
        c = ws.cell(row=row, column=2, value=sachet_weight)
        style_cell(c, total_font, total_fill, Alignment(horizontal='center'), thin_border)
        c.number_format = '0.0'
        c = ws.cell(row=row, column=3, value=1)
        style_cell(c, total_font, total_fill, Alignment(horizontal='center'), thin_border)
        c.number_format = '0.0%'
        row += 2

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14

def create_night_soak_sheet(ws):
    ws.title = 'Night Soak'
    ws.sheet_properties.tabColor = '2E4057'
    sachet_weight = 44

    ws.merge_cells('A1:C1')
    c = ws['A1']
    c.value = 'Night Soak - Daily Recipe (44g per sachet)'
    style_cell(c, Font(name='Arial', bold=True, color='FFFFFF', size=14), header_fill, Alignment(horizontal='center'))
    for col in range(1, 4):
        style_cell(ws.cell(row=1, column=col), fill=header_fill, border=thin_border)

    ws.merge_cells('A3:C3')
    c = ws.cell(row=3, column=1, value='Same recipe every day of the week')
    style_cell(c, Font(name='Arial', italic=True, size=10, color='666666'), align=Alignment(horizontal='center'))

    row = 5
    for label, col_idx in [('Ingredient', 1), ('Grams', 2), ('Percentage', 3)]:
        c = ws.cell(row=row, column=col_idx, value=label)
        style_cell(c, sub_header_font, sub_header_fill, Alignment(horizontal='center'), thin_border)
    row += 1

    for name, grams in night_soak:
        ws.cell(row=row, column=1, value=name).font = data_font
        ws.cell(row=row, column=2, value=grams).font = data_font
        ws.cell(row=row, column=2).number_format = '0.0'
        pct = round(grams / sachet_weight * 100, 1)
        ws.cell(row=row, column=3, value=pct / 100).font = data_font
        ws.cell(row=row, column=3).number_format = '0.0%'
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center' if col > 1 else 'left')
        row += 1

    c = ws.cell(row=row, column=1, value='Total')
    style_cell(c, total_font, total_fill, border=thin_border)
    c = ws.cell(row=row, column=2, value=sachet_weight)
    style_cell(c, total_font, total_fill, Alignment(horizontal='center'), thin_border)
    c.number_format = '0.0'
    c = ws.cell(row=row, column=3, value=1)
    style_cell(c, total_font, total_fill, Alignment(horizontal='center'), thin_border)
    c.number_format = '0.0%'

    # Weekly summary
    row += 3
    ws.merge_cells(f'A{row}:C{row}')
    c = ws.cell(row=row, column=1, value='Weekly Requirement (7 sachets)')
    style_cell(c, day_header_font, day_header_fill, Alignment(horizontal='center'))
    for col in range(1, 4):
        style_cell(ws.cell(row=row, column=col), fill=day_header_fill, border=thin_border)
    row += 1

    for label, col_idx in [('Ingredient', 1), ('Per Sachet (g)', 2), ('Weekly Total (g)', 3)]:
        c = ws.cell(row=row, column=col_idx, value=label)
        style_cell(c, sub_header_font, sub_header_fill, Alignment(horizontal='center'), thin_border)
    row += 1

    for name, grams in night_soak:
        ws.cell(row=row, column=1, value=name).font = data_font
        ws.cell(row=row, column=2, value=grams).font = data_font
        ws.cell(row=row, column=2).number_format = '0.0'
        ws.cell(row=row, column=3, value=grams * 7).font = data_font
        ws.cell(row=row, column=3).number_format = '0.0'
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center' if col > 1 else 'left')
        row += 1

    c = ws.cell(row=row, column=1, value='Total')
    style_cell(c, total_font, total_fill, border=thin_border)
    c = ws.cell(row=row, column=2, value=sachet_weight)
    style_cell(c, total_font, total_fill, Alignment(horizontal='center'), thin_border)
    c.number_format = '0.0'
    c = ws.cell(row=row, column=3, value=sachet_weight * 7)
    style_cell(c, total_font, total_fill, Alignment(horizontal='center'), thin_border)
    c.number_format = '0.0'

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 18

create_day_pack_sheet(wb.active)
create_night_soak_sheet(wb.create_sheet())

output_path = 'C:/Users/Thira/OneDrive/Documents/GitHub/wkly-nuts/WKLY_Nuts_Recipes.xlsx'
wb.save(output_path)
print(f'Saved to {output_path}')
